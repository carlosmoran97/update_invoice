# -*- coding: utf-8 -*-
"""
Author: Carlos Morán
Description: this script generates the postgresql code necesary to update the correlatives of invoices in Odoo ERP.
"""

from openpyxl import load_workbook

filepath = input("Filepath: ")

wb = load_workbook(filepath)

sheet = wb.get_sheet_by_name(input('Sheet: '))

max_row = sheet.max_row
max_column = sheet.max_column

initial_row = int(input("Initial row: "))
initial_column = int(input("Initial column: "))

for i in range(initial_row, max_row + 1):

    for j in range(initial_column, max_column + 1):

        # print(cell_obj.value, end=' | ')
        # escribir la sentencia sql
        if j == max_column:
            cell_obj = sheet.cell(row=i, column=j-1)
            if cell_obj.value != None:
				# correlativo de la factura en físico
                cell_obj_fisico = sheet.cell(row=i, column=j-2)
				# fecha de la factura
                cell_obj_fecha = sheet.cell(row=i, column=j)
                sql = """
                        BEGIN TRANSACTION;
                        UPDATE account_invoice SET number='{}' WHERE number='{}' AND date_invoice='{}';
                        COMMIT TRANSACTION;
                      """.format(
                    cell_obj.value, cell_obj_fisico.value, cell_obj_fecha.value)
                print(sql)
